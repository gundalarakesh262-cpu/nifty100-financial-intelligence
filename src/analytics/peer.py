{
  "nbformat": 4,
  "nbformat_minor": 0,
  "metadata": {
    "colab": {
      "provenance": [],
      "authorship_tag": "ABX9TyM6zeYgo4ozWdssBus8EcjJ",
      "include_colab_link": true
    },
    "kernelspec": {
      "name": "python3",
      "display_name": "Python 3"
    },
    "language_info": {
      "name": "python"
    }
<<<<<<< HEAD

    out_rows: List[Dict[str, Optional[float]]] = []
    required_cols = [col for _, (col, _) in metrics.items()]
    if peer_group_col not in df.columns:
        return pd.DataFrame(out_rows)

    base = df.copy()
    base['peer_group_name'] = base[peer_group_col]

    for group_name, group in base.groupby('peer_group_name'):
        if pd.isna(group_name):
            continue
        for _, row in group.iterrows():
            for metric_name, (col, higher) in metrics.items():
                if col not in group.columns:
                    value = None
                    rank = None
                else:
                    value = row[col]
                    rank = percentile_rank(group[col], higher).loc[row.name]
                out_rows.append({
                    'company_id': row.get('company_id'),
                    'peer_group_name': group_name,
                    'metric': metric_name,
                    'value': value,
                    'percentile_rank': rank,
                    'year': row.get('year'),
                })

    return pd.DataFrame(out_rows)


def write_peer_percentiles_sqlite(df: pd.DataFrame, db_path: str = 'nifty100.db') -> None:
    if df.empty:
        return
    path = Path(db_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(str(path)) as conn:
        df.to_sql('peer_percentiles', conn, if_exists='replace', index=False)


def generate_peer_comparison_excel(
    df: pd.DataFrame,
    peer_groups: pd.DataFrame,
    output_path: str = 'output/peer_comparison.xlsx',
) -> None:
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    if 'company_id' in df.columns:
        df['company_id'] = df['company_id'].astype(str).str.strip().str.upper()
    merged = attach_peer_groups(df, peer_groups)

    metrics = [
        'return_on_equity_pct',
        'return_on_capital_employed_pct',
        'net_profit_margin_pct',
        'debt_to_equity',
        'free_cash_flow_cr',
        'pat_cagr_5y_pct',
        'revenue_cagr_5y_pct',
        'eps_cagr_5y_pct',
        'interest_coverage',
        'asset_turnover',
    ]

    ranks = merged.copy()
    # compute percentiles per peer group but ensure the result aligns with ranks.index
    for metric in metrics:
        if metric in ranks.columns:
            percentile_series = pd.Series(index=ranks.index, dtype=float)
            for group_name, group in ranks.groupby('peer_group_name', dropna=False):
                idx = group.index
                series = pd.to_numeric(group[metric], errors='coerce')
                if series.dropna().empty:
                    percentile_series.loc[idx] = 50.0
                else:
                    ranks_pct = series.rank(pct=True, method='max')
                    if metric != 'debt_to_equity':
                        percentile_series.loc[idx] = ranks_pct.fillna(0) * 100.0
                    else:
                        percentile_series.loc[idx] = (1 - ranks_pct).fillna(0) * 100.0
            ranks[f'{metric}_percentile'] = percentile_series.fillna(50.0)
        else:
            ranks[f'{metric}_percentile'] = None

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for group_name, group in ranks.groupby('peer_group_name'):
            if pd.isna(group_name):
                continue
            sheet_name = str(group_name)[:31]
            cols = [
                'company_id',
                'company_name',
            ] + metrics + [f'{metric}_percentile' for metric in metrics]
            subset = group[[c for c in cols if c in group.columns]]
            medians = subset.median(numeric_only=True)
            subset.loc['Peer Group Median'] = medians
            subset.to_excel(writer, sheet_name=sheet_name, index=False)

    # Apply conditional formatting and highlight benchmark row after saving
    workbook = load_workbook(output_file)
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    gold_fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
    bold_font = Font(bold=True)

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        percentile_cols = [i + 1 for i, h in enumerate(headers) if isinstance(h, str) and h.endswith('_percentile')]
        company_col = next((i + 1 for i, h in enumerate(headers) if h == 'company_id'), None)
        benchmark_ids = set(peer_groups.loc[peer_groups['peer_group_name'] == sheet_name, 'company_id'].astype(str))

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            company_id = str(row[company_col - 1].value).strip() if company_col else ''
            is_benchmark = company_id in benchmark_ids
            for col_idx in percentile_cols:
                cell = row[col_idx - 1]
                val = cell.value
                try:
                    pct = float(val)
                except Exception:
                    continue
                if pct >= 75:
                    cell.fill = green_fill
                elif pct <= 25:
                    cell.fill = red_fill
                else:
                    cell.fill = yellow_fill
            if is_benchmark:
                for cell in row:
                    cell.fill = gold_fill
                    cell.font = bold_font

        for col_idx in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col_idx)].auto_size = True

    workbook.save(output_file)


def generate_radar_charts(
    df: pd.DataFrame,
    peer_groups: pd.DataFrame,
    output_dir: str = 'reports/radar_charts',
) -> None:
    import matplotlib.pyplot as plt
    from matplotlib.ticker import MaxNLocator

    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    radar_metrics = [
        ('ROE', 'return_on_equity_pct'),
        ('ROCE', 'return_on_capital_employed_pct'),
        ('NPM', 'net_profit_margin_pct'),
        ('D/E', 'debt_to_equity'),
        ('FCF', 'free_cash_flow_cr'),
        ('PAT CAGR 5yr', 'pat_cagr_5y_pct'),
        ('Revenue CAGR 5yr', 'revenue_cagr_5y_pct'),
        ('Composite Score', 'composite_quality_score'),
    ]

    merged = attach_peer_groups(df, peer_groups)
    merged['company_id'] = merged['company_id'].astype(str).str.strip().str.upper()

    gauges = len(radar_metrics)
    angles = [n / float(gauges) * 2 * np.pi for n in range(gauges)]
    angles += angles[:1]

    for group_name, group in merged.groupby('peer_group_name'):
        if pd.isna(group_name):
            continue
        peer_avg = {}
        for label, col in radar_metrics:
            if col in group.columns:
                if label == 'D/E':
                    peer_avg[label] = 100.0 - percentile_rank(group[col], False).mean()
                else:
                    peer_avg[label] = np.nanmean(pd.to_numeric(group[col], errors='coerce'))
            else:
                peer_avg[label] = 0.0

        peer_values = [peer_avg[label] for label, _ in radar_metrics]
        peer_values += peer_values[:1]

        for _, row in group.iterrows():
            company_values = []
            for label, col in radar_metrics:
                company_values.append(float(row[col]) if pd.notna(row.get(col)) else 0.0)
            company_values += company_values[:1]

            plt.figure(figsize=(8, 8))
            ax = plt.subplot(111, polar=True)
            plt.xticks(angles[:-1], [label for label, _ in radar_metrics], color='grey', size=8)
            ax.set_rlabel_position(30)
            ax.plot(angles, peer_values, linewidth=1, linestyle='dashed', label='Peer Avg')
            ax.fill(angles, peer_values, alpha=0.1)
            ax.plot(angles, company_values, linewidth=2, linestyle='solid', label=str(row.get('company_id', '')))
            ax.fill(angles, company_values, alpha=0.25)
            plt.title(f"{row.get('company_id', '')} — {group_name}", size=11, y=1.1)
            plt.legend(loc='upper right', bbox_to_anchor=(1.3, 1.1))
            plt.tight_layout()
            filename = output_path / f"{row.get('company_id', '')}_radar.png"
            plt.savefig(filename, dpi=150)
            plt.close()

    if merged['peer_group_name'].isna().any():
        no_group = merged[merged['peer_group_name'].isna()]
        avg = merged[merged['peer_group_name'].notna()]
        if not avg.empty:
            nifty_avg = {}
            for label, col in radar_metrics:
                nifty_avg[label] = np.nanmean(pd.to_numeric(avg.get(col), errors='coerce'))
            values = [float(no_group.iloc[0].get(col, 0.0)) for _, col in radar_metrics]
            values += values[:1]
            avg_values = [float(nifty_avg[label]) for label, _ in radar_metrics]
            avg_values += avg_values[:1]
            plt.figure(figsize=(8, 8))
            ax = plt.subplot(111, polar=True)
            plt.xticks(angles[:-1], [label for label, _ in radar_metrics], color='grey', size=8)
            ax.plot(angles, avg_values, linewidth=1, linestyle='dashed', label='Nifty 100 Avg')
            ax.plot(angles, values, linewidth=2, linestyle='solid', label=str(no_group.iloc[0].get('company_id', '')))
            ax.fill(angles, values, alpha=0.25)
            plt.title(f"{no_group.iloc[0].get('company_id', '')} — No Peer Group", size=11, y=1.1)
            plt.legend(loc='upper right', bbox_to_anchor=(1.3, 1.1))
            plt.tight_layout()
            filename = output_path / f"{no_group.iloc[0].get('company_id', '')}_radar.png"
            plt.savefig(filename, dpi=150)
            plt.close()
=======
  },
  "cells": [
    {
      "cell_type": "markdown",
      "metadata": {
        "id": "view-in-github",
        "colab_type": "text"
      },
      "source": [
        "<a href=\"https://colab.research.google.com/github/gundalarakesh262-cpu/nifty100-financial-intelligence/blob/main/src/analytics/peer.py\" target=\"_parent\"><img src=\"https://colab.research.google.com/assets/colab-badge.svg\" alt=\"Open In Colab\"/></a>"
      ]
    },
    {
      "cell_type": "code",
      "source": [
        "from openpyxl import load_workbook\n",
        "\n",
        "excel_path = \"output/peer_comparison.xlsx\"\n",
        "\n",
        "wb = load_workbook(excel_path)\n",
        "\n",
        "print(\"Total sheets:\", len(wb.sheetnames))\n",
        "print(\"Sheet names:\", wb.sheetnames)\n",
        "\n",
        "for sheet_name in wb.sheetnames:\n",
        "    ws = wb[sheet_name]\n",
        "    headers = [cell.value for cell in ws[1]]\n",
        "\n",
        "    percentile_cols = [\n",
        "        h for h in headers\n",
        "        if isinstance(h, str) and h.endswith(\"_percentile\")\n",
        "    ]\n",
        "\n",
        "    normal_metric_cols = [\n",
        "        h for h in headers\n",
        "        if h not in [\"company_id\", \"company_name\"]\n",
        "        and isinstance(h, str)\n",
        "        and not h.endswith(\"_percentile\")\n",
        "    ]\n",
        "\n",
        "    last_row_values = [cell.value for cell in ws[ws.max_row]]\n",
        "\n",
        "    print(\"\\nSheet:\", sheet_name)\n",
        "    print(\"Rows:\", ws.max_row)\n",
        "    print(\"Columns:\", ws.max_column)\n",
        "    print(\"Metric columns:\", len(normal_metric_cols))\n",
        "    print(\"Percentile columns:\", len(percentile_cols))\n",
        "    print(\"Last row starts with:\", last_row_values[:2])"
      ],
      "metadata": {
        "colab": {
          "base_uri": "https://localhost:8080/"
        },
        "id": "qnAu2Py1ZQRq",
        "outputId": "f92a6445-2134-4234-da78-6fc1ebec3896"
      },
      "execution_count": 8,
      "outputs": [
        {
          "output_type": "stream",
          "name": "stdout",
          "text": [
            "Total sheets: 11\n",
            "Sheet names: ['Automobiles', 'Consumer Finance', 'FMCG', 'IT Services', 'Life Insurance', 'Oil & Gas', 'Pharmaceuticals', 'Power & Utilities', 'Private Banks', 'Public Sector Banks', 'Steel']\n",
            "\n",
            "Sheet: Automobiles\n",
            "Rows: 9\n",
            "Columns: 20\n",
            "Metric columns: 8\n",
            "Percentile columns: 10\n",
            "Last row starts with: [None, None]\n",
            "\n",
            "Sheet: Consumer Finance\n",
            "Rows: 5\n",
            "Columns: 20\n",
            "Metric columns: 8\n",
            "Percentile columns: 10\n",
            "Last row starts with: [None, None]\n",
            "\n",
            "Sheet: FMCG\n",
            "Rows: 9\n",
            "Columns: 20\n",
            "Metric columns: 8\n",
            "Percentile columns: 10\n",
            "Last row starts with: [None, None]\n",
            "\n",
            "Sheet: IT Services\n",
            "Rows: 7\n",
            "Columns: 20\n",
            "Metric columns: 8\n",
            "Percentile columns: 10\n",
            "Last row starts with: [None, None]\n",
            "\n",
            "Sheet: Life Insurance\n",
            "Rows: 6\n",
            "Columns: 20\n",
            "Metric columns: 8\n",
            "Percentile columns: 10\n",
            "Last row starts with: [None, None]\n",
            "\n",
            "Sheet: Oil & Gas\n",
            "Rows: 7\n",
            "Columns: 20\n",
            "Metric columns: 8\n",
            "Percentile columns: 10\n",
            "Last row starts with: [None, None]\n",
            "\n",
            "Sheet: Pharmaceuticals\n",
            "Rows: 7\n",
            "Columns: 20\n",
            "Metric columns: 8\n",
            "Percentile columns: 10\n",
            "Last row starts with: [None, None]\n",
            "\n",
            "Sheet: Power & Utilities\n",
            "Rows: 9\n",
            "Columns: 20\n",
            "Metric columns: 8\n",
            "Percentile columns: 10\n",
            "Last row starts with: [None, None]\n",
            "\n",
            "Sheet: Private Banks\n",
            "Rows: 7\n",
            "Columns: 20\n",
            "Metric columns: 8\n",
            "Percentile columns: 10\n",
            "Last row starts with: [None, None]\n",
            "\n",
            "Sheet: Public Sector Banks\n",
            "Rows: 6\n",
            "Columns: 20\n",
            "Metric columns: 8\n",
            "Percentile columns: 10\n",
            "Last row starts with: [None, None]\n",
            "\n",
            "Sheet: Steel\n",
            "Rows: 6\n",
            "Columns: 20\n",
            "Metric columns: 8\n",
            "Percentile columns: 10\n",
            "Last row starts with: [None, None]\n"
          ]
        }
      ]
    },
    {
      "cell_type": "code",
      "source": [
        "from peer import load_peer_groups, generate_peer_comparison_excel\n",
        "import pandas as pd\n",
        "from google.colab import files\n",
        "\n",
        "ratios = pd.read_csv(\"financial_ratios_generated.csv\")\n",
        "peer_groups = load_peer_groups(\"peer_groups_cleaned.csv\")\n",
        "\n",
        "ratios[\"fiscal_year\"] = pd.to_numeric(ratios[\"fiscal_year\"], errors=\"coerce\")\n",
        "\n",
        "latest = (\n",
        "    ratios\n",
        "    .dropna(subset=[\"fiscal_year\"])\n",
        "    .sort_values([\"company_id\", \"fiscal_year\"])\n",
        "    .groupby(\"company_id\")\n",
        "    .tail(1)\n",
        "    .reset_index(drop=True)\n",
        ")\n",
        "\n",
        "generate_peer_comparison_excel(\n",
        "    latest,\n",
        "    peer_groups,\n",
        "    output_path=\"output/peer_comparison.xlsx\"\n",
        ")\n",
        "\n",
        "files.download(\"output/peer_comparison.xlsx\")"
      ],
      "metadata": {
        "colab": {
          "base_uri": "https://localhost:8080/",
          "height": 453
        },
        "id": "oVPqVMBuc4yA",
        "outputId": "d27f0747-c971-4328-98a3-7fd2f2bf77bc"
      },
      "execution_count": 9,
      "outputs": [
        {
          "output_type": "error",
          "ename": "ModuleNotFoundError",
          "evalue": "No module named 'peer'",
          "traceback": [
            "\u001b[0;31m---------------------------------------------------------------------------\u001b[0m",
            "\u001b[0;31mModuleNotFoundError\u001b[0m                       Traceback (most recent call last)",
            "\u001b[0;32m/tmp/ipykernel_1418/3961797134.py\u001b[0m in \u001b[0;36m<cell line: 0>\u001b[0;34m()\u001b[0m\n\u001b[0;32m----> 1\u001b[0;31m \u001b[0;32mfrom\u001b[0m \u001b[0mpeer\u001b[0m \u001b[0;32mimport\u001b[0m \u001b[0mload_peer_groups\u001b[0m\u001b[0;34m,\u001b[0m \u001b[0mgenerate_peer_comparison_excel\u001b[0m\u001b[0;34m\u001b[0m\u001b[0;34m\u001b[0m\u001b[0m\n\u001b[0m\u001b[1;32m      2\u001b[0m \u001b[0;32mimport\u001b[0m \u001b[0mpandas\u001b[0m \u001b[0;32mas\u001b[0m \u001b[0mpd\u001b[0m\u001b[0;34m\u001b[0m\u001b[0;34m\u001b[0m\u001b[0m\n\u001b[1;32m      3\u001b[0m \u001b[0;32mfrom\u001b[0m \u001b[0mgoogle\u001b[0m\u001b[0;34m.\u001b[0m\u001b[0mcolab\u001b[0m \u001b[0;32mimport\u001b[0m \u001b[0mfiles\u001b[0m\u001b[0;34m\u001b[0m\u001b[0;34m\u001b[0m\u001b[0m\n\u001b[1;32m      4\u001b[0m \u001b[0;34m\u001b[0m\u001b[0m\n\u001b[1;32m      5\u001b[0m \u001b[0mratios\u001b[0m \u001b[0;34m=\u001b[0m \u001b[0mpd\u001b[0m\u001b[0;34m.\u001b[0m\u001b[0mread_csv\u001b[0m\u001b[0;34m(\u001b[0m\u001b[0;34m\"financial_ratios_generated.csv\"\u001b[0m\u001b[0;34m)\u001b[0m\u001b[0;34m\u001b[0m\u001b[0;34m\u001b[0m\u001b[0m\n",
            "\u001b[0;31mModuleNotFoundError\u001b[0m: No module named 'peer'",
            "",
            "\u001b[0;31m---------------------------------------------------------------------------\u001b[0;32m\nNOTE: If your import is failing due to a missing package, you can\nmanually install dependencies using either !pip or !apt.\n\nTo view examples of installing some common dependencies, click the\n\"Open Examples\" button below.\n\u001b[0;31m---------------------------------------------------------------------------\u001b[0m\n"
          ],
          "errorDetails": {
            "actions": [
              {
                "action": "open_url",
                "actionText": "Open Examples",
                "url": "/notebooks/snippets/importing_libraries.ipynb"
              }
            ]
          }
        }
      ]
    },
    {
      "cell_type": "code",
      "source": [],
      "metadata": {
        "id": "ly0fCiXafDBt"
      },
      "execution_count": null,
      "outputs": []
    }
  ]
}
>>>>>>> 8152794c14e7af76fae7490aad79e5add51e59bd
