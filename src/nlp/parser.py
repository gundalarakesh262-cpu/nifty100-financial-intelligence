"""Parse analysis workbook text fields into normalized NLP outputs."""

import re
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_WORKBOOK_PATH = ROOT / "data" / "raw" / "analysis.xlsx"
DEFAULT_OUTPUT_DIR = ROOT / "output"
DEFAULT_PROCESSED_DIR = ROOT / "data" / "processed"

TARGET_FIELDS = (
    "compounded_sales_growth",
    "compounded_profit_growth",
    "stock_price_cagr",
    "roe",
)

PRIMARY_PATTERN = re.compile(r"(\d+)\s*Years?:?\s*([\d.]+)%", re.IGNORECASE)
SIGNED_PATTERN = re.compile(r"(\d+)\s*Years?:?\s*(-?[\d.]+)%", re.IGNORECASE)


if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from src.analytics.cagr import compute_cagr
from src.etl.normaliser import normalize_ticker, normalize_year


def parse_metric_text(text: object) -> Optional[Tuple[int, float]]:
    """Parse a metric text cell like '10 Years: 21%' into period and value."""
    if text is None or pd.isna(text):
        return None

    text_str = str(text).strip()
    if not text_str:
        return None

    match = PRIMARY_PATTERN.search(text_str)
    if match:
        return int(match.group(1)), float(match.group(2))

    match = SIGNED_PATTERN.search(text_str)
    if match:
        return int(match.group(1)), float(match.group(2))

    return None


def _normalize_company_id(value: object) -> Optional[str]:
    if value is None or pd.isna(value):
        return None

    company_id = normalize_ticker(value)
    if company_id is None:
        return None

    company_id = company_id.strip()
    return company_id or None


def _parse_period_date(value: object) -> Optional[pd.Timestamp]:
    if value is None or pd.isna(value):
        return None

    year_str = str(value).strip()
    if not year_str:
        return None

    normalized = normalize_year(year_str)
    if normalized:
        try:
            year_part, month_part = normalized.split("-")
            return pd.Timestamp(year=int(year_part), month=int(month_part), day=1)
        except Exception:
            return None

    match = re.match(r"^(\d{4})$", year_str)
    if match:
        return pd.Timestamp(year=int(match.group(1)), month=3, day=1)

    return None


def _load_analysis_sheet(workbook_path: Path) -> Tuple[List[str], List[Tuple[int, List[object]]]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    header_row = [cell.value for cell in next(worksheet.iter_rows(min_row=2, max_row=2))]
    headers = [str(value).strip() if value is not None else "" for value in header_row]

    rows: List[Tuple[int, List[object]]] = []
    for row_number, row in enumerate(worksheet.iter_rows(min_row=3, values_only=True), start=3):
        rows.append((row_number, list(row)))

    return headers, rows


def _as_path(value: object) -> Path:
    if isinstance(value, Path):
        return value
    return Path(value)


def build_parse_records(workbook_path: Path = DEFAULT_WORKBOOK_PATH) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Return parsed metric rows and parse failures from the analysis workbook."""
    workbook_path = _as_path(workbook_path)

    if not workbook_path.exists():
        raise FileNotFoundError(f"Analysis workbook not found: {workbook_path}")

    headers, rows = _load_analysis_sheet(workbook_path)
    column_index = {name: index for index, name in enumerate(headers) if name}

    required_columns = ["company_id", *TARGET_FIELDS]
    missing_columns = [column for column in required_columns if column not in column_index]
    if missing_columns:
        raise ValueError(f"Workbook is missing required columns: {missing_columns}")

    parsed_rows: List[Dict[str, object]] = []
    failure_rows: List[Dict[str, object]] = []

    for row_number, row_values in rows:
        company_id = _normalize_company_id(row_values[column_index["company_id"]])
        if not company_id:
            continue

        for metric_type in TARGET_FIELDS:
            raw_value = row_values[column_index[metric_type]]
            parsed = parse_metric_text(raw_value)

            if parsed is None:
                failure_rows.append(
                    {
                        "company_id": company_id,
                        "metric_type": metric_type,
                        "raw_text": None if raw_value is None else str(raw_value),
                        "row_number": row_number,
                        "failure_reason": "PATTERN_MISMATCH"
                        if raw_value not in (None, "")
                        else "EMPTY_VALUE",
                    }
                )
                continue

            period_years, value_pct = parsed
            parsed_rows.append(
                {
                    "company_id": company_id,
                    "metric_type": metric_type,
                    "period_years": period_years,
                    "value_pct": value_pct,
                }
            )

    parsed_df = pd.DataFrame(parsed_rows)
    failures_df = pd.DataFrame(failure_rows)
    return parsed_df, failures_df


def _load_yearly_series(
    csv_path: Path,
    value_column: str,
    year_column: str = "year",
    date_column: Optional[str] = None,
) -> Dict[str, pd.DataFrame]:
    if not csv_path.exists():
        return {}

    frame = pd.read_csv(csv_path)
    if "company_id" not in frame.columns or value_column not in frame.columns:
        return {}

    frame["company_id"] = frame["company_id"].astype(str).str.strip().str.upper()

    if date_column is not None and date_column in frame.columns:
        frame["period_date"] = pd.to_datetime(frame[date_column], errors="coerce")
        frame = frame.dropna(subset=["period_date", value_column])
        frame = frame.sort_values(["company_id", "period_date"])
        frame["period_date"] = frame["period_date"].dt.to_period("Y").dt.to_timestamp()
    elif year_column in frame.columns:
        frame["period_date"] = frame[year_column].apply(_parse_period_date)
        frame = frame.dropna(subset=["period_date", value_column])
        frame = frame.sort_values(["company_id", "period_date"])
    else:
        return {}

    frame = frame.groupby(["company_id", "period_date"], as_index=False)[value_column].last()

    grouped: Dict[str, pd.DataFrame] = {}
    for company_id, company_frame in frame.groupby("company_id"):
        grouped[company_id] = company_frame.sort_values("period_date").reset_index(drop=True)

    return grouped


def _compute_series_cagr(series_frame: pd.DataFrame, value_column: str, period_years: int) -> Tuple[Optional[float], Optional[str]]:
    if series_frame is None or series_frame.empty:
        return None, "NO_COMPUTED_SOURCE"

    if len(series_frame) < period_years + 1:
        return None, "INSUFFICIENT_HISTORY"

    start_value = series_frame.iloc[-(period_years + 1)][value_column]
    end_value = series_frame.iloc[-1][value_column]
    return compute_cagr(start_value, end_value, period_years)


def build_validation_records(
    parsed_df: pd.DataFrame,
    processed_dir: Path = DEFAULT_PROCESSED_DIR,
) -> pd.DataFrame:
    """Cross-validate parsed metrics against computed company history."""
    processed_dir = _as_path(processed_dir)

    if parsed_df.empty:
        return pd.DataFrame(
            columns=
            [
                "company_id",
                "metric_type",
                "period_years",
                "parsed_value_pct",
                "computed_value_pct",
                "delta_pct",
                "validation_status",
                "review_required",
            ]
        )

    sources = {
        "compounded_sales_growth": _load_yearly_series(
            processed_dir / "profitandloss_cleaned.csv",
            "sales",
        ),
        "compounded_profit_growth": _load_yearly_series(
            processed_dir / "profitandloss_cleaned.csv",
            "net_profit",
        ),
        "stock_price_cagr": _load_yearly_series(
            processed_dir / "stock_prices_cleaned.csv",
            "adjusted_close",
            year_column="date",
            date_column="date",
        ),
        "roe": _load_yearly_series(
            processed_dir / "financial_ratios_generated.csv",
            "return_on_equity_pct",
        ),
    }

    validation_rows: List[Dict[str, object]] = []

    for row in parsed_df.itertuples(index=False):
        series_map = sources.get(row.metric_type, {})
        series_frame = series_map.get(row.company_id)

        if series_frame is None or series_frame.empty:
            validation_rows.append(
                {
                    "company_id": row.company_id,
                    "metric_type": row.metric_type,
                    "period_years": row.period_years,
                    "parsed_value_pct": row.value_pct,
                    "computed_value_pct": None,
                    "delta_pct": None,
                    "validation_status": "NO_COMPUTED_SOURCE",
                    "review_required": True,
                }
            )
            continue

        computed_value, computed_flag = _compute_series_cagr(
            series_frame,
            series_frame.columns[-1],
            int(row.period_years),
        )

        if computed_value is None:
            validation_rows.append(
                {
                    "company_id": row.company_id,
                    "metric_type": row.metric_type,
                    "period_years": row.period_years,
                    "parsed_value_pct": row.value_pct,
                    "computed_value_pct": None,
                    "delta_pct": None,
                    "validation_status": computed_flag or "INSUFFICIENT_HISTORY",
                    "review_required": True,
                }
            )
            continue

        delta_pct = abs(float(row.value_pct) - float(computed_value))
        validation_rows.append(
            {
                "company_id": row.company_id,
                "metric_type": row.metric_type,
                "period_years": row.period_years,
                "parsed_value_pct": float(row.value_pct),
                "computed_value_pct": float(computed_value),
                "delta_pct": delta_pct,
                "validation_status": "DIVERGENCE" if delta_pct > 5.0 else "MATCH",
                "review_required": delta_pct > 5.0,
            }
        )

    return pd.DataFrame(validation_rows)


def write_outputs(
    parsed_df: pd.DataFrame,
    failures_df: pd.DataFrame,
    validation_df: pd.DataFrame,
    output_dir: Path = DEFAULT_OUTPUT_DIR,
) -> Tuple[Path, Path, Path]:
    output_dir = _as_path(output_dir)

    output_dir.mkdir(parents=True, exist_ok=True)

    parsed_path = output_dir / "analysis_parsed.csv"
    failures_path = output_dir / "parse_failures.csv"
    validation_path = output_dir / "analysis_validation.csv"

    parsed_df.to_csv(parsed_path, index=False)
    failures_df.to_csv(failures_path, index=False)
    validation_df.to_csv(validation_path, index=False)

    return parsed_path, failures_path, validation_path


def main(
    workbook_path: Path = DEFAULT_WORKBOOK_PATH,
    output_dir: Path = DEFAULT_OUTPUT_DIR,
    processed_dir: Path = DEFAULT_PROCESSED_DIR,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    workbook_path = _as_path(workbook_path)
    output_dir = _as_path(output_dir)
    processed_dir = _as_path(processed_dir)

    parsed_df, failures_df = build_parse_records(workbook_path)
    validation_df = build_validation_records(parsed_df, processed_dir)
    write_outputs(parsed_df, failures_df, validation_df, output_dir)
    return parsed_df, failures_df, validation_df


if __name__ == "__main__":
    main()
