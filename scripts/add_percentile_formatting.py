from pathlib import Path
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule

base = Path(__file__).resolve().parent.parent
input_xlsx = base / 'output' / 'peer_comparison.xlsx'
output_xlsx = base / 'output' / 'peer_comparison_percentile_formatted.xlsx'

if not input_xlsx.exists():
    print('Missing', input_xlsx)
    raise SystemExit(1)

wb = load_workbook(input_xlsx)
# color scale: red (low) -> yellow (mid) -> green (high)
rule = ColorScaleRule(start_type='min', start_color='F8696B',
                      mid_type='percentile', mid_value=50, mid_color='FFEB84',
                      end_type='max', end_color='63BE7B')

for ws in wb.worksheets:
    if ws.max_row < 2:
        continue
    # read header row
    headers = [cell.value for cell in ws[1]]
    if not headers:
        continue
    # find percentile-like columns
    percentile_cols = []
    for idx, h in enumerate(headers, start=1):
        if not h:
            continue
        key = str(h).lower()
        if 'peer_score' in key or 'peer_composite' in key or key.endswith('_pct') or 'percentile' in key:
            percentile_cols.append(idx)
    # apply color scale to each column's data range (from row 2 to max_row)
    for col_idx in percentile_cols:
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        data_range = f"{col_letter}2:{col_letter}{ws.max_row}"
        ws.conditional_formatting.add(data_range, rule)

wb.save(output_xlsx)
print('Wrote formatted workbook:', output_xlsx)
